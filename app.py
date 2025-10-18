from flask import Flask, render_template, request, jsonify
import threading
import io
import pandas as pd
from openpyxl import load_workbook
import os

app = Flask(__name__)

# Global storage
processed_data = []
COLUMNS_DISPLAY = [
    'Account No', 'DATE', 'TRANSACTION DETAILS', 'VALUE DATE',
    'WITHDRAWAL AMT', 'DEPOSIT AMT', 'BALANCE AMT'
]


# Upload & processing
@app.route("/", methods=["GET", "POST"])
def upload_file():
    global processed_data
    if request.method == "POST":
        file = request.files.get('file')
        if file:
            filename = file.filename
            processed_data = []

            # Read file into memory to avoid "closed file" error
            file_content = file.read()
            file_stream = io.BytesIO(file_content)

            # Start background processing
            if filename.endswith('.csv'):
                threading.Thread(target=process_csv, args=(file_stream,)).start()
            elif filename.endswith(('.xls', '.xlsx')):
                threading.Thread(target=process_large_excel, args=(file_stream,)).start()
            else:
                processed_data = [{'error': 'Unsupported file type'}]

            return '', 202  # Accepted

    return render_template("upload.html")


# CSV processing
def process_csv(file_stream):
    global processed_data
    processed_data = []

    chunksize = 5000
    reader = pd.read_csv(file_stream, chunksize=chunksize)
    normalized_to_display = {col.strip().replace(" ", "").lower(): col for col in COLUMNS_DISPLAY}

    for chunk in reader:
        chunk.columns = [col.strip().replace(" ", "").lower() for col in chunk.columns]

        if 'withdrawalamt' in chunk.columns:
            chunk['fraud_flag'] = pd.to_numeric(chunk['withdrawalamt'], errors='coerce').fillna(0) > 10000
        else:
            chunk['fraud_flag'] = False

        display_cols = [norm for norm in normalized_to_display.keys() if norm in chunk.columns]
        chunk_dict = chunk[display_cols + ['fraud_flag']].rename(
            columns={norm: display for norm, display in normalized_to_display.items() if norm in chunk.columns}
        ).to_dict(orient='records')

        processed_data.extend(chunk_dict)


# Large Excel processing (>80k rows)
def process_large_excel(file_stream):
    global processed_data
    processed_data = []

    wb = load_workbook(file_stream, read_only=True)
    ws = wb.active

    header = [str(cell).strip().replace(" ", "").lower() for cell in next(ws.iter_rows(values_only=True))]
    normalized_to_display = {col.strip().replace(" ", "").lower(): col for col in COLUMNS_DISPLAY}

    chunk = []
    chunk_size = 5000

    for row in ws.iter_rows(values_only=True):
        row_dict = dict(zip(header, row))
        withdrawal = float(row_dict.get('withdrawalamt') or 0)
        row_dict['fraud_flag'] = withdrawal > 10000

        display_row = {normalized_to_display.get(k, k): v for k, v in row_dict.items() if k in normalized_to_display}
        display_row['fraud_flag'] = row_dict['fraud_flag']
        chunk.append(display_row)

        if len(chunk) >= chunk_size:
            processed_data.extend(chunk)
            chunk = []

    if chunk:
        processed_data.extend(chunk)


# Endpoint: Check if data is ready
@app.route("/is_ready")
def is_ready():
    return jsonify({'ready': len(processed_data) > 0})


# Fetch processed data (paginated)
@app.route("/get_data")
def get_data():
    start = int(request.args.get('start', 0))
    limit = int(request.args.get('limit', 5000))
    end = start + limit
    return jsonify(processed_data[start:end])


# New route: Serve sample data CSV
@app.route("/sample_data")
def sample_data():
    global processed_data
    processed_data = []

    # ✅ Correct path: static/data/sample_transactions_long.csv
    sample_path = os.path.join(app.static_folder, 'data', 'sample_transactions_long.csv')

    print(f"Looking for sample file at: {sample_path}")  # debug print

    if not os.path.exists(sample_path):
        print("❌ File not found")
        return jsonify({'error': f'Sample file not found at {sample_path}'}), 404

    try:
        with open(sample_path, 'rb') as f:
            file_stream = io.BytesIO(f.read())
            process_csv(file_stream)

        print(f"✅ Loaded {len(processed_data)} rows from sample data.")
        return jsonify(processed_data[:5000])  # send first chunk
    except Exception as e:
        print(f"⚠️ Error loading sample data: {e}")
        return jsonify({'error': str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True)
